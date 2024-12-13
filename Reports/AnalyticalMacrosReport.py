import time
import logging
from datetime import datetime, timedelta
import schedule
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from openpyxl import load_workbook
import os

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


def transfer_data_to_historic(download_folder):
    downloaded_file_path = os.path.join(download_folder, "macroAnalitico.rpt.xls")
    historic_file_path = r"/Relatório Analítico de Macros.xlsx"

    if os.path.exists(downloaded_file_path):
        df = pd.read_excel(downloaded_file_path)

        if "Fim do relatório" in df.iloc[-1].to_string():
            df = df.iloc[:-1]

        if os.path.exists(historic_file_path):
            wb = load_workbook(historic_file_path)
            sheet = wb.active
            last_row = sheet.max_row + 1
        else:
            wb = load_workbook(historic_file_path)
            sheet = wb.active
            last_row = 1

        for index, row in df.iterrows():
            sheet.append(row.values.tolist())

        wb.save(historic_file_path)
        logging.info("Dados transferidos para o arquivo de histórico com sucesso.")

        os.remove(downloaded_file_path)
        logging.info(f"Arquivo baixado {downloaded_file_path} removido após processamento.")
    else:
        logging.error(f"Arquivo {downloaded_file_path} não encontrado.")


def login():
    chrome_options = Options()
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--remote-debugging-port=9222")

    service = Service(ChromeDriverManager(driver_version="131.0.6778.108").install())
    driver = webdriver.Chrome(service=service, options=chrome_options)

    logging.info("Abrindo a página de login")
    driver.get("https://")

    try:
        enterprise_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "empresa"))
        )
        user_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "usuario"))
        )
        password_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "senha"))
        )

        enterprise_field.send_keys("")
        user_field.send_keys("")
        password_field.send_keys("")

        enter_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "btnLogar"))
        )
        enter_button.click()

        report_section = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//span[contains(@class, 'menu-label') and text()='Relatórios']"))
        )
        report_section.click()

        report_section_submenu = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//a[@data-pagina-id='paginaRelatorios']"))
        )
        print("Cliquei uma vez no submenu de Relatórios")
        report_section_submenu.click()
        time.sleep(10)
        iframe = driver.find_element(By.ID, "paginaRelatorios")
        driver.switch_to.frame(iframe)

        codigo_input = driver.find_element(By.ID, "fCodigo")
        codigo_input.send_keys("17")

        consultar_btn = driver.find_element(By.ID, "btnConsultar")
        consultar_btn.click()

        time.sleep(5)

        plus_btn = driver.find_element(By.XPATH, "//*[@id='linha_010200000000']/td[1]/a/span")
        plus_btn.click()

        time.sleep(5)

        list_btn = driver.find_element(By.CLASS_NAME, "icon-list-alt")
        list_btn.click()

        time.sleep(10)

        yesterday = datetime.now() - timedelta(days=1)
        date_yesterday = yesterday.strftime("%d/%m/%Y")

        date_init = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='fPar_3']"))
        )
        date_init.send_keys(date_yesterday)

        time.sleep(2)

        date_end = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='fPar_4']"))
        )
        date_end.send_keys(date_yesterday)

        time.sleep(5)

        xls_btn = driver.find_element(By.ID, "btnXLS")
        xls_btn.click()

        driver.switch_to.default_content()

        time.sleep(15)

        download_folder = os.path.expanduser("~/Downloads")

        transfer_data_to_historic(download_folder)

    except Exception as e:
        logging.error(f"Erro: {e}")
    except TimeoutError:
        logging.error("Erro: Tempo de espera excedido em algum elemento.")
    finally:
        time.sleep(15)
        driver.quit()


login()
